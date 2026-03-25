import math
import io
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
#  PAGE CONFIG
# ──────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="DU Calculation Sheet",
    page_icon="🔧",
    layout="wide",
)

# ──────────────────────────────────────────────────────────────────────────────
#  CONSTANTS  — exact values from the uploaded Excel sheet
# ──────────────────────────────────────────────────────────────────────────────

KITCHEN_FIXTURES = [
    {"sr": 1, "name": "KITCHEN SINK",                           "du": 0.8},
    {"sr": 2, "name": "WM (WASHING MACHINE UP TO 6 KG CONSIDER)", "du": 0.8},
    {"sr": 3, "name": "DISH WASHER",                            "du": 0.8},
]

TOILET_FIXTURES = [
    {"sr": 1, "name": "W C- WITH 6 L CISTERN",                  "du": 2.0},
    {"sr": 2, "name": "FLOOR TRAP- Shower without plug consider","du": 0.6},
    {"sr": 3, "name": "WASHBASIN",                               "du": 0.5},
    {"sr": 4, "name": "Trap",                                    "du": 1.5},
    {"sr": 5, "name": "WM (WASHING MACHINE UP TO 6 KG CONSIDER)","du": 0.8},
]

# Pipe size lookup  ─  (max_flow_ls, label)
# Flow formula from sheet:  TOTAL FLOW = 0.5 × √(TOTAL DU)
PIPE_SIZES = [
    (0.50,  "DN 50"),
    (0.90,  "DN 63"),
    (1.20,  "DN 75"),
    (2.50,  "DN 100"),
    (5.20,  "DN 110  With Swept Tee  5.2"),
    (7.60,  "DN 125  With Swept Tee  7.6"),
    (12.40, "DN 160"),
    (22.00, "DN 200"),
]


# ──────────────────────────────────────────────────────────────────────────────
#  CALCULATION HELPERS  — mirror the Excel formulas exactly
# ──────────────────────────────────────────────────────────────────────────────

def for_toilet(qty: float, du: float) -> float:
    """E = C × D  →  QTY × DU"""
    return qty * du


def total_du_row(floors: float, ft: float) -> float:
    """G = F × E  →  FLOORS × FOR_TOILET"""
    return floors * ft


def total_flow(sum_du: float) -> float:
    """H = 0.5 × √(G_total)"""
    if sum_du <= 0:
        return 0.0
    return 0.5 * math.sqrt(sum_du)


def req_dia(flow: float) -> str:
    for max_f, label in PIPE_SIZES:
        if flow <= max_f:
            return label
    return f"DN 200+  ({flow:.2f} l/s — exceeds table)"


# ──────────────────────────────────────────────────────────────────────────────
#  CSS
# ──────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* ── page background ── */
[data-testid="stAppViewContainer"] { background: #f5f7fa; }

/* ── section title banner ── */
.sec-title {
    background: linear-gradient(90deg,#1a3c6e,#2e6da4);
    color:#fff; font-weight:700; font-size:1.05rem;
    padding:.55rem 1.1rem; border-radius:7px; margin-bottom:.6rem;
    letter-spacing:.03em;
}

/* ── result card ── */
.res-card {
    background:#fff; border:1px solid #cde;
    border-left:5px solid #2e6da4;
    border-radius:8px; padding:.9rem 1.3rem; margin-top:.5rem;
}

/* ── metric strip ── */
.metric-strip {
    display:flex; gap:1rem; flex-wrap:wrap; margin:.4rem 0;
}
.m-box {
    flex:1; min-width:130px;
    background:#fff; border:1px solid #dde;
    border-radius:8px; padding:.6rem 1rem; text-align:center;
}
.m-box .label { font-size:.72rem; color:#666; text-transform:uppercase; }
.m-box .val   { font-size:1.4rem; font-weight:700; color:#1a3c6e; }

/* ── table header ── */
thead tr th {
    background:#1a3c6e !important;
    color:#fff !important;
    text-align:center !important;
}
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────────────────────
#  HEADER
# ──────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div style="background:linear-gradient(90deg,#1a3c6e,#2e6da4);
            color:#fff;padding:1.2rem 2rem;border-radius:10px;margin-bottom:1.2rem;">
  <h2 style="margin:0">🔧 DU Calculation Sheet</h2>
  <p  style="margin:0;opacity:.85;font-size:.9rem">
      Drainage Unit Calculator &nbsp;|&nbsp; Formula: TOTAL FLOW = 0.5 × √(ΣTOTAL DU)
  </p>
</div>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────────────────────
#  SECTION RENDERER
# ──────────────────────────────────────────────────────────────────────────────

def render_section(title: str, fixtures: list, key_prefix: str):
    """
    Render one section (Kitchen or Toilet).
    Returns (rows_list, sum_total_du, flow, dia_string).
    """
    st.markdown(f'<div class="sec-title">{title}</div>', unsafe_allow_html=True)

    # ── column headers (match Excel exactly) ──
    hdr = st.columns([0.5, 3.5, 1.2, 0.8, 1.2, 1.2, 1.2, 1.4, 2.0])
    for h, label in zip(hdr, [
        "SR NO.", "FIXTURES", "QTY ✏️", "D U",
        "FOR TOILET", "FLOORS ✏️", "TOTAL DU",
        "TOTAL FLOW\n(l/s)", "REQ DIA (mm)"
    ]):
        h.markdown(
            f"<div style='font-size:.72rem;font-weight:700;color:#1a3c6e;"
            f"text-align:center;border-bottom:2px solid #2e6da4;"
            f"padding-bottom:.2rem'>{label}</div>",
            unsafe_allow_html=True,
        )

    rows = []
    sum_du = 0.0

    for fx in fixtures:
        cols = st.columns([0.5, 3.5, 1.2, 0.8, 1.2, 1.2, 1.2, 1.4, 2.0])

        qty    = cols[2].number_input(" ", min_value=0, max_value=200, value=0,
                                      step=1, key=f"{key_prefix}_qty_{fx['sr']}",
                                      label_visibility="collapsed")
        floors = cols[5].number_input(" ", min_value=0, max_value=100, value=0,
                                      step=1, key=f"{key_prefix}_flr_{fx['sr']}",
                                      label_visibility="collapsed")

        ft  = for_toilet(qty, fx["du"])   # E = C × D
        tdu = total_du_row(floors, ft)     # G = F × E
        sum_du += tdu

        # display cells
        cols[0].markdown(f"<p style='text-align:center;margin:0;padding:.45rem 0'>{fx['sr']}</p>", unsafe_allow_html=True)
        cols[1].markdown(f"<p style='margin:0;padding:.45rem 0;font-size:.85rem'>{fx['name']}</p>", unsafe_allow_html=True)
        cols[3].markdown(f"<p style='text-align:center;margin:0;padding:.45rem 0'>{fx['du']}</p>", unsafe_allow_html=True)
        cols[4].markdown(f"<p style='text-align:center;margin:0;padding:.45rem 0'>{ft:.2f}</p>", unsafe_allow_html=True)
        cols[6].markdown(f"<p style='text-align:center;margin:0;padding:.45rem 0'><b>{tdu:.2f}</b></p>", unsafe_allow_html=True)

        rows.append({
            "SR NO.":           fx["sr"],
            "FIXTURES":         fx["name"],
            "QTY":              qty,
            "D U":              fx["du"],
            "FOR TOILET":       round(ft, 4),
            "FLOORS":           floors,
            "TOTAL DU":         round(tdu, 4),
        })

    flow = total_flow(sum_du)
    dia  = req_dia(flow)

    # ── totals row ──
    st.markdown("<hr style='margin:.3rem 0;border-color:#cde'>", unsafe_allow_html=True)
    tcols = st.columns([0.5, 3.5, 1.2, 0.8, 1.2, 1.2, 1.2, 1.4, 2.0])
    tcols[1].markdown("<b style='font-size:.85rem'>TOTAL</b>", unsafe_allow_html=True)
    tcols[6].markdown(f"<b style='color:#1a3c6e'>{sum_du:.2f}</b>", unsafe_allow_html=True)
    tcols[7].markdown(f"<b style='color:#1a3c6e'>{flow:.3f}</b>", unsafe_allow_html=True)
    tcols[8].markdown(f"<span style='font-size:.8rem;color:#c00'><b>({dia})</b></span>", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    return rows, round(sum_du, 4), round(flow, 4), dia


# ──────────────────────────────────────────────────────────────────────────────
#  MAIN LAYOUT  —  two sections side-by-side on wide screens
# ──────────────────────────────────────────────────────────────────────────────
left, right = st.columns(2, gap="large")

with left:
    k_rows, k_sum_du, k_flow, k_dia = render_section(
        "TYPICAL KITCHEN", KITCHEN_FIXTURES, "k"
    )

with right:
    t_rows, t_sum_du, t_flow, t_dia = render_section(
        "Typical Toilet", TOILET_FIXTURES, "t"
    )

# ──────────────────────────────────────────────────────────────────────────────
#  SUMMARY STRIP
# ──────────────────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown("### 📊 Summary")

c1, c2, c3, c4, c5, c6 = st.columns(6)
def mbox(col, label, val):
    col.markdown(
        f"""<div class="m-box">
              <div class="label">{label}</div>
              <div class="val">{val}</div>
            </div>""",
        unsafe_allow_html=True,
    )

mbox(c1, "Kitchen Total DU",    f"{k_sum_du:.2f}")
mbox(c2, "Kitchen Flow (l/s)",  f"{k_flow:.3f}")
mbox(c3, "Kitchen Req Dia",     k_dia.split()[0] + " " + (k_dia.split()[1] if len(k_dia.split())>1 else ""))
mbox(c4, "Toilet Total DU",     f"{t_sum_du:.2f}")
mbox(c5, "Toilet Flow (l/s)",   f"{t_flow:.3f}")
mbox(c6, "Toilet Req Dia",      t_dia.split()[0] + " " + (t_dia.split()[1] if len(t_dia.split())>1 else ""))

# ──────────────────────────────────────────────────────────────────────────────
#  FORMULA NOTE
# ──────────────────────────────────────────────────────────────────────────────
with st.expander("ℹ️  Formula reference (mirrors the Excel sheet exactly)", expanded=False):
    st.markdown("""
| Column | Formula | Description |
|--------|---------|-------------|
| **FOR TOILET** | `= QTY × D U` | Discharge units for one floor |
| **TOTAL DU**   | `= FLOORS × FOR TOILET` | Total discharge units per fixture |
| **TOTAL DU (sum)** | `= SUM(TOTAL DU)` | Sum of all fixtures |
| **TOTAL FLOW (l/s)** | `= 0.5 × √(ΣTOTAL DU)` | EN 12056-2, K = 0.5 |
| **REQ DIA (mm)** | Pipe-size lookup on flow | DN 50 → DN 200 |

**Fixture DU values (from sheet):**

| Section | Fixture | DU |
|---------|---------|-----|
| Kitchen | Kitchen Sink | 0.8 |
| Kitchen | Washing Machine (≤6 kg) | 0.8 |
| Kitchen | Dish Washer | 0.8 |
| Toilet | WC – 6 L Cistern | 2.0 |
| Toilet | Floor Trap / Shower (no plug) | 0.6 |
| Toilet | Washbasin | 0.5 |
| Toilet | Trap | 1.5 |
| Toilet | Washing Machine (≤6 kg) | 0.8 |
""")

# ──────────────────────────────────────────────────────────────────────────────
#  EXCEL EXPORT  — reproduces the original sheet layout
# ──────────────────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown("### 📥 Export")

def build_excel(k_rows, k_sum_du, k_flow, k_dia,
                t_rows, t_sum_du, t_flow, t_dia) -> bytes:

    wb = Workbook()
    ws = wb.active
    ws.title = "DU Calculation Sheet"

    # ── styles ──
    hdr_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill   = PatternFill("solid", fgColor="1A3C6E")
    sec_font   = Font(name="Arial", bold=True, size=11)
    sec_fill   = PatternFill("solid", fgColor="2E6DA4")
    sec_font_w = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    tot_fill   = PatternFill("solid", fgColor="D9E8F5")
    tot_font   = Font(name="Arial", bold=True, size=10)
    cell_font  = Font(name="Arial", size=10)
    red_font   = Font(name="Arial", bold=True, color="C00000", size=10)
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    thin  = Side(style="thin",   color="AABBCC")
    thick = Side(style="medium", color="1A3C6E")
    border_all  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
    border_bot  = Border(left=thin,  right=thin,  top=thick, bottom=thick)

    COLS = ["A","B","C","D","E","F","G","H","I"]
    HEADERS = ["SR NO.","FIXTURES","QTY","D U","FOR TOILET","FLOORS",
               "TOTAL DU","TOTAL FLOW( l/s)","REQ DIA (mm)"]
    COL_W   = [7, 42, 8, 7, 13, 9, 12, 18, 28]

    for i, (col, w) in enumerate(zip(COLS, COL_W)):
        ws.column_dimensions[col].width = w

    row = 1   # current Excel row pointer

    def write_section(title, fixtures_data, sum_du, flow, dia, data_rows):
        nonlocal row

        # section title row
        ws.merge_cells(f"A{row}:I{row}")
        c = ws[f"A{row}"]
        c.value = title
        c.font  = sec_font_w
        c.fill  = PatternFill("solid", fgColor="1A3C6E")
        c.alignment = left_al
        ws.row_dimensions[row].height = 20
        row += 1

        # header row
        for col_l, h in zip(COLS, HEADERS):
            c = ws[f"{col_l}{row}"]
            c.value     = h
            c.font      = hdr_font
            c.fill      = PatternFill("solid", fgColor="2E6DA4")
            c.alignment = center
            c.border    = border_all
        ws.row_dimensions[row].height = 30
        row += 1

        # data rows  — use formulas exactly as original sheet
        data_start = row
        for dr in data_rows:
            ws[f"A{row}"] = dr["SR NO."]
            ws[f"B{row}"] = dr["FIXTURES"]
            ws[f"C{row}"] = dr["QTY"]         # USER INPUT
            ws[f"D{row}"] = dr["D U"]
            ws[f"E{row}"] = f"=(C{row}*D{row})"          # FOR TOILET
            ws[f"F{row}"] = dr["FLOORS"]      # USER INPUT
            ws[f"G{row}"] = f"=F{row}*E{row}"             # TOTAL DU

            for col_l in COLS:
                c = ws[f"{col_l}{row}"]
                c.font      = cell_font
                c.border    = border_all
                c.alignment = center if col_l != "B" else left_al

            ws[f"C{row}"].fill = PatternFill("solid", fgColor="FFF2CC")  # yellow = input
            ws[f"F{row}"].fill = PatternFill("solid", fgColor="FFF2CC")  # yellow = input
            ws.row_dimensions[row].height = 18
            row += 1

        data_end = row - 1

        # totals row
        ws[f"B{row}"]  = "TOTAL"
        ws[f"G{row}"]  = f"=SUM(G{data_start}:G{data_end})"
        ws[f"H{row}"]  = f"=0.5*SQRT(G{row})"
        ws[f"I{row}"]  = f"({dia})"

        for col_l in COLS:
            c = ws[f"{col_l}{row}"]
            c.font      = tot_font
            c.fill      = tot_fill
            c.border    = border_bot
            c.alignment = center if col_l != "B" else left_al

        ws[f"I{row}"].font = red_font
        ws.row_dimensions[row].height = 20
        row += 2   # blank separator

    write_section("TYPICAL KITCHEN", KITCHEN_FIXTURES, k_sum_du, k_flow, k_dia, k_rows)
    write_section("Typical Toilet",  TOILET_FIXTURES,  t_sum_du, t_flow, t_dia, t_rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


if st.button("📄 Generate Excel Report", type="primary"):
    xls = build_excel(k_rows, k_sum_du, k_flow, k_dia,
                      t_rows, t_sum_du, t_flow, t_dia)
    st.download_button(
        label="⬇️  Download  DU_Calculation_Sheet.xlsx",
        data=xls,
        file_name="DU_Calculation_Sheet.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
